import openpyxl as opxl

import os
import re
import shutil
import csv
import json


init_dict = {
    'path_xlsx_folder': 'C:/Users/ArthurEmig/Documents/Private_documents/0000_idea/recipes.xlsx',
    'worksheet_lebensmittel': 'ingreds',
    'worksheets_recipes': ['r_day_1', 'r_day_2', 'r_day_3'],
    'name_csv_ingreds': 'Lebensmittel.csv',
    'name_json_recipes': 'recipes.json',
}


workbook_all = opxl.load_workbook(init_dict['path_xlsx_folder'])


# ---- get worksheets with data ----
worksheet_ingreds = workbook_all[init_dict['worksheet_lebensmittel']]
worksheets_recipes = [workbook_all[i] for i in init_dict['worksheets_recipes']]

# ----------------- get the ingreds csv-----------------
ingred_list_ids = []
ingred_list_names = []
ingred_list_amount = []
ingred_list_units = []
ingred_list_mhd = []

start_row = 2
end_row = worksheet_ingreds.max_row

for row_num in range(start_row, end_row + 1):
    if worksheet_ingreds.cell(row=row_num, column=1).value is None:
        break
    ingred_list_ids.append(worksheet_ingreds.cell(row=row_num, column=1).value)
    ingred_list_names.append(worksheet_ingreds.cell(row=row_num, column=2).value)
    ingred_list_mhd.append(worksheet_ingreds.cell(row=row_num, column=3).value)
    ingred_list_amount.append(worksheet_ingreds.cell(row=row_num, column=9).value)
    ingred_list_units.append(worksheet_ingreds.cell(row=row_num, column=10).value)

pass

# create a dictionary 'ingred_name' <-> 'ingred_id'
ingred_name_to_id = {ingred_name: ingred_id for ingred_name, ingred_id in zip(ingred_list_names, ingred_list_ids)}


rows = zip(ingred_list_ids, ingred_list_names, ingred_list_mhd, ingred_list_amount, ingred_list_units)

with open(init_dict['name_csv_ingreds'], "w", newline='') as f:
    writer = csv.writer(f)
    for row in rows:
        writer.writerow(row)

# ----------------- get recipes -----------------
start_row = 1
none_rows = 0

recipes = []
recipe = {}
# initialize lists of ingredient features included in a particular recipe
ingred_ids_list, ingred_names_list, ingred_amounts_list = [], [], []

for ws_num, ws_recipes in enumerate(worksheets_recipes):

    # ------- iterate over all rows in the worksheet with recipe info -------
    end_row = ws_recipes.max_row
    # initialize the content of previous first column cell
    # which is to be memorized for worksheet data parsing
    prev_first_col_cell = ''
    prev_second_col_cell = ''
    prev_third_col_cell = ''
    prev_meal_type = ''
    for row_num in range(start_row, end_row+1):

        # initialize a dictionary of a particular recipe
        # get the value of
        first_column_cell_value = ws_recipes.cell(row=row_num, column=1).value
        second_column_cell_value = ws_recipes.cell(row=row_num, column=2).value
        third_column_cell_value = ws_recipes.cell(row=row_num, column=3).value

        if second_column_cell_value is None and third_column_cell_value is None and row_num != 0:

        # if second_column_cell_value is None and third_column_cell_value is None and (ws_num != 0 and row_num != 0):
            recipes.append(recipe)
            # reset the recipe dict
            recipe = {}
            # reset lists of ingredients
            ingred_ids_list, ingred_names_list, ingred_amounts_list = [], [], []

        if first_column_cell_value == 'breakfast':
            # meal type
            meal_type = 0
            # memorize the meal type
            prev_meal_type = meal_type
            recipe.update({'meal_type': meal_type})
        if first_column_cell_value == 'lunch':
            # meal type
            meal_type = 1
            # memorize the meal type
            prev_meal_type = meal_type
            recipe.update({'meal_type': meal_type})
        if first_column_cell_value == 'dinner':
            meal_type = 2
            # memorize the meal type
            prev_meal_type = meal_type
            recipe.update({'meal_type': meal_type})
        if prev_first_col_cell in ['breakfast', 'lunch', 'dinner']:
            recipe.update({'name': first_column_cell_value})

        if third_column_cell_value is None and not(second_column_cell_value is None):
            recipe.update({'name': second_column_cell_value})
            # reset lists of ingredients
            ingred_ids_list, ingred_names_list, ingred_amounts_list = [], [], []

        if first_column_cell_value is None and not(prev_second_col_cell is None) and not(third_column_cell_value is None):
            recipe.update({'meal_type': prev_meal_type})
            ingred_names_list.append(second_column_cell_value)
            ingred_ids_list.append(ingred_name_to_id[second_column_cell_value])
            ingred_amounts_list.append(ws_recipes.cell(row=row_num, column=3).value)

            recipe.update({
                'ingred_ids': ingred_ids_list,
                'ingred_amounts': ingred_amounts_list,
                'ingred_names': ingred_names_list
            })
        prev_first_col_cell = ws_recipes.cell(row=row_num, column=1).value
        prev_second_col_cell = ws_recipes.cell(row=row_num, column=2).value
        prev_third_col_cell = ws_recipes.cell(row=row_num, column=3).value
pass
final_recipes = []
for num_recipe, recipe_item in enumerate(recipes[1:]):
    recipe_item.update({'recipe_id': num_recipe})
    final_recipes.append(recipe_item)


with open(init_dict['name_json_recipes'], 'w') as fout:
    json.dump(final_recipes, fout)

